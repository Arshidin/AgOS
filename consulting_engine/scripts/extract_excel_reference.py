"""Extract reference values from Excel template for test validation.

Usage: python3 scripts/extract_excel_reference.py
Output: tests/fixtures/excel_reference.json

Layout of Operating Model sheet:
  Cols 10-33 (J-AG): Monthly data — 24 months (Aug 2026 — Jul 2028)
  Col 34: gap
  Cols 35-46 (AI-AT): Annual data — 12 years (2026-2037, year-end values)
  Row 39: dates (EOMONTH)
"""

import json
import openpyxl
import datetime
import sys
from pathlib import Path

EXCEL_PATH = Path(__file__).parent.parent.parent / "Docs" / "Zengi.Farm_Model farm_020426_v10_WintSumm.xlsx"
OUTPUT_PATH = Path(__file__).parent.parent / "tests" / "fixtures" / "excel_reference.json"

# Operating Model column ranges
MONTHLY_COL_START = 10  # Column J = month 1
MONTHLY_COL_END = 33    # Column AG = month 24
ANNUAL_COL_START = 35   # Column AI = year 1
ANNUAL_COL_END = 46     # Column AT = year 12


def safe_float(v):
    """Convert to float, handling None and non-numeric."""
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def read_monthly(ws, row):
    """Read 24 monthly values from Operating Model."""
    return [safe_float(ws.cell(row=row, column=c).value)
            for c in range(MONTHLY_COL_START, MONTHLY_COL_END + 1)]


def read_annual(ws, row):
    """Read 12 annual values from Operating Model."""
    return [safe_float(ws.cell(row=row, column=c).value)
            for c in range(ANNUAL_COL_START, ANNUAL_COL_END + 1)]


def extract_timeline(ws):
    """Timeline dates from Operating Model row 39."""
    monthly_dates = []
    for col in range(MONTHLY_COL_START, MONTHLY_COL_END + 1):
        v = ws.cell(row=39, column=col).value
        if isinstance(v, datetime.datetime):
            monthly_dates.append(v.strftime("%Y-%m-%d"))
        else:
            monthly_dates.append(str(v) if v else None)

    annual_dates = []
    for col in range(ANNUAL_COL_START, ANNUAL_COL_END + 1):
        v = ws.cell(row=39, column=col).value
        if isinstance(v, datetime.datetime):
            annual_dates.append(v.strftime("%Y-%m-%d"))
        else:
            annual_dates.append(str(v) if v else None)

    return {
        "monthly_dates": monthly_dates,
        "annual_dates": annual_dates,
        "monthly_count": len(monthly_dates),
        "annual_count": len(annual_dates),
        "start_date": monthly_dates[0] if monthly_dates else None,
    }


def extract_input(wb):
    """Input parameters from Input sheet + Operating Model header."""
    ws_input = wb["Input"]
    ws_om = wb["Operating Model"]

    params = {}
    # Input sheet: column A = label, column F = value
    for row in range(1, ws_input.max_row + 1):
        for label_col, val_col in [(1, 6), (1, 2), (3, 4), (5, 6)]:
            label = ws_input.cell(row=row, column=label_col).value
            value = ws_input.cell(row=row, column=val_col).value
            if label and value is not None:
                key = str(label).strip()
                if key not in params:
                    params[key] = value

    # Key Operating Model params
    om_params = {}
    for row in range(5, 49):
        label = ws_om.cell(row=row, column=2).value
        val = ws_om.cell(row=row, column=5).value or ws_om.cell(row=row, column=10).value
        if label:
            om_params[str(label).strip()] = val

    return {"input_sheet": params, "operating_model_header": om_params}


def extract_herd(ws):
    """Herd turnover — 6 groups from Operating Model.

    Rows (verified from Excel inspection):
    50-58: Cows (маточное поголовье)
    60-67: Bulls (быки-производители)
    69-77: Calves (приплод)
    79-88: Heifers (тёлки)
    90-97: Steers (бычки)
    162-170: Summary
    """
    return {
        "monthly_count": MONTHLY_COL_END - MONTHLY_COL_START + 1,
        "cows": {
            "bop": read_monthly(ws, 50),
            "purchased": read_monthly(ws, 51),
            "from_heifers": read_monthly(ws, 52),
            "culled": read_monthly(ws, 53),
            "mortality": read_monthly(ws, 54),
            "interim": read_monthly(ws, 55),
            "sold_breeding": read_monthly(ws, 56),
            "eop": read_monthly(ws, 57),
            "avg": read_monthly(ws, 58),
        },
        "cows_annual_eop": read_annual(ws, 57),
        "bulls": {
            "bop": read_monthly(ws, 60),
            "purchased": read_monthly(ws, 61),
            "from_steers": read_monthly(ws, 62),
            "culled": read_monthly(ws, 63),
            "mortality": read_monthly(ws, 64),
            "eop": read_monthly(ws, 66),
            "avg": read_monthly(ws, 67),
        },
        "bulls_annual_eop": read_annual(ws, 66),
        "calves": {
            "bop": read_monthly(ws, 69),
            "born": read_monthly(ws, 70),
            "before_split": read_monthly(ws, 71),
            "to_heifers": read_monthly(ws, 72),
            "to_steers": read_monthly(ws, 73),
            "eop": read_monthly(ws, 76),
            "avg": read_monthly(ws, 77),
        },
        "heifers": {
            "bop": read_monthly(ws, 79),
            "from_calves": read_monthly(ws, 80),
            "mortality": read_monthly(ws, 81),
            "before": read_monthly(ws, 82),
            "sold_breeding": read_monthly(ws, 83),
            "to_cows": read_monthly(ws, 85),
            "eop": read_monthly(ws, 87),
            "avg": read_monthly(ws, 88),
        },
        "steers": {
            "bop": read_monthly(ws, 90),
            "from_calves": read_monthly(ws, 91),
            "to_bulls": read_monthly(ws, 92),
            "mortality": read_monthly(ws, 93),
            "sold": read_monthly(ws, 94),
            "eop": read_monthly(ws, 96),
            "avg": read_monthly(ws, 97),
        },
        "total_avg_livestock": read_monthly(ws, 102),
        "total_eop": read_monthly(ws, 105),
        "total_avg_annual": read_annual(ws, 102),
        "sold_heifers_breeding": read_monthly(ws, 165),
        "sold_cows_culled": read_monthly(ws, 166),
        "sold_bulls_culled": read_monthly(ws, 167),
        "sold_own_steers": read_monthly(ws, 168),
        "total_sold": read_monthly(ws, 170),
    }


def extract_capex(wb):
    """CAPEX items from CAPEX sheet."""
    ws = wb["CAPEX"]
    items = []
    for row in range(5, ws.max_row + 1):
        code = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=2).value
        cost = ws.cell(row=row, column=8).value
        if name and cost is not None:
            items.append({
                "row": row,
                "code": str(code) if code else "",
                "name": str(name).strip(),
                "cost": safe_float(cost),
            })
    return {"items": items}


def extract_staff(wb):
    """Staff payroll from Staff sheet."""
    ws = wb["Staff"]

    # Scan for positions and total
    positions = []
    total_payroll = 0.0
    for row in range(1, ws.max_row + 1):
        b = ws.cell(row=row, column=2).value
        j = ws.cell(row=row, column=10).value
        if b and j is not None:
            name = str(b).strip()
            val = safe_float(j)
            if "итого" in name.lower() or "total" in name.lower():
                total_payroll = val
            else:
                positions.append({"row": row, "name": name, "col_j": val})

    return {"positions": positions, "total_monthly_payroll": total_payroll}


def extract_wacc(wb):
    """WACC parameters from WACC sheet (labels in col C, values in col E)."""
    ws = wb["WACC"]
    values = {}
    for row in range(3, ws.max_row + 1):
        label = ws.cell(row=row, column=3).value  # Column C
        val = ws.cell(row=row, column=5).value     # Column E
        if label and val is not None and isinstance(val, (int, float)):
            values[str(label).strip()] = val
    return values


def extract_pnl(ws):
    """P&L from Operating Model."""
    return {
        "revenue_monthly": read_monthly(ws, 188),
        "revenue_annual": read_annual(ws, 188),
        "gross_profit_monthly": read_monthly(ws, 222),
        "ebitda_monthly": read_monthly(ws, 228),
        "net_profit_monthly": read_monthly(ws, 240),
        "net_profit_annual": read_annual(ws, 240),
    }


def extract_cashflow(ws):
    """Cash flow from Operating Model."""
    return {
        "cash_eop_monthly": read_monthly(ws, 348),
        "cash_eop_annual": read_annual(ws, 348),
    }


def extract_feeding(wb):
    """Feeding costs from Cattle Feeding Cycle sheet."""
    ws = wb["Cattle Feeding Cycle"]
    # Row 247 is total reproducer feeding cost
    # Check if same column layout
    vals = []
    for col in range(10, 50):
        v = ws.cell(row=247, column=col).value
        if v is not None:
            vals.append(safe_float(v))
    return {
        "total_reproducer_row247": vals,
    }


def main():
    print(f"Reading: {EXCEL_PATH}")
    if not EXCEL_PATH.exists():
        print(f"ERROR: File not found: {EXCEL_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws_om = wb["Operating Model"]

    reference = {}

    print("Extracting timeline...")
    reference["timeline"] = extract_timeline(ws_om)

    print("Extracting input parameters...")
    reference["input"] = extract_input(wb)

    print("Extracting herd turnover...")
    reference["herd"] = extract_herd(ws_om)

    print("Extracting CAPEX...")
    reference["capex"] = extract_capex(wb)

    print("Extracting staff...")
    reference["staff"] = extract_staff(wb)

    print("Extracting WACC...")
    reference["wacc"] = extract_wacc(wb)

    print("Extracting P&L...")
    reference["pnl"] = extract_pnl(ws_om)

    print("Extracting cash flow...")
    reference["cashflow"] = extract_cashflow(ws_om)

    print("Extracting feeding...")
    reference["feeding"] = extract_feeding(wb)

    # Write output
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(reference, f, ensure_ascii=False, indent=2, default=str)

    print(f"\nOutput: {OUTPUT_PATH}")
    print(f"Size: {OUTPUT_PATH.stat().st_size / 1024:.1f} KB")

    # Quick validation summary
    herd = reference["herd"]
    print(f"\n=== Validation Summary ===")
    print(f"Monthly data: {herd['monthly_count']} months")
    print(f"Timeline start: {reference['timeline']['start_date']}")
    print(f"Cows EOP month 1: {herd['cows']['eop'][0]}")
    print(f"Cows EOP month 24: {herd['cows']['eop'][-1]}")
    print(f"Cows annual EOP: {herd['cows_annual_eop'][:3]}...")
    print(f"Bulls EOP month 1: {herd['bulls']['eop'][0]}")
    print(f"Staff payroll: {reference['staff']['total_monthly_payroll']}")
    print(f"WACC keys: {list(reference['wacc'].keys())[:5]}...")


if __name__ == "__main__":
    main()
